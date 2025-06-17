import os
import scipy.io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import logging

# --- Configuration & Setup ---

# Get the absolute path of the script
script_path = os.path.abspath(__file__)
# Get the directory containing the script
script_dir = os.path.dirname(script_path)

# Construct paths relative to the script's directory
directory = os.path.join(script_dir, "../server/data")
output_excel = os.path.join(script_dir, "experiment_summary.xlsx")
log_file_path = os.path.join(script_dir, "formatData.log")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_file_path),
        logging.StreamHandler()
    ]
)

logging.info("Script started.")
logging.info(f"Script path: {script_path}")
logging.info(f"Data directory: {directory}")
logging.info(f"Output Excel file: {output_excel}")
logging.info(f"Log file: {log_file_path}")

# Default max_number, can be overridden by command-line argument
default_max_number = 4000

# --- Argument Parsing ---
import argparse
parser = argparse.ArgumentParser(description="Process .mat experiment files and generate an Excel summary.")
parser.add_argument(
    "--max_files",
    type=int,
    default=default_max_number,
    help=f"Maximum number of experiment files (exp-*.mat) to check (default: {default_max_number})"
)
args = parser.parse_args()
max_number_to_process = args.max_files
logging.info(f"Will process up to {max_number_to_process} experiment files based on --max_files argument.")


# Gather data grouped by (year, func)
grouped_data = {}

logging.info(f"Starting data processing loop for files up to exp-{max_number_to_process}.mat...")
for i in range(1, max_number_to_process + 1):
    file_name = f"exp-{i}.mat"
    file_path = os.path.join(directory, file_name)
    logging.info(f"Attempting to process file: {file_path}")

    if not os.path.exists(file_path):
        logging.warning(f"File not found: {file_path}. Skipping.")
        continue

    try:
        mat = scipy.io.loadmat(file_path)
        data = mat['data']

        # Flatten structured array if necessary
        if isinstance(data, np.ndarray) and data.dtype.names:
            data = data[0, 0]

        # Critical data extraction with error handling
        try:
            year = str(data['year'][0])
            func = int(data['func'][0][0])
            repeats = int(data['repeat'][0][0])
            selection_methods = str(data['selectionMethods']) if data['adaptive'] == 0 else "Adaptive"
            final_fitness = data['finalFitness']
        except (KeyError, IndexError, TypeError) as e:
            logging.error(f"Error extracting critical data from {file_name}: {e}. Skipping this file.")
            continue

    except (KeyError, AttributeError, Exception) as e:
        logging.error(f"Error loading or parsing {file_name}: {e}. Skipping this file.")
        continue

    # Extract the best fitness value from each repetition
    best_fitness_per_repeat = []
    logging.info(f"Processing {repeats} repeats for {file_name} (Year: {year}, Func: {func})")
    for j in range(repeats):
        logging.info(f"Processing repeat {j+1}/{repeats} for {file_name}")
        try:
            # Ensure final_fitness[0][j] is valid before trying to flatten
            if final_fitness is not None and \
               isinstance(final_fitness, np.ndarray) and \
               final_fitness.ndim >= 2 and \
               final_fitness.shape[0] > 0 and \
               final_fitness.shape[1] > j and \
               final_fitness[0][j] is not None:
                
                fitness_values_raw = final_fitness[0][j]
                # Check if it's a scalar or needs flattening
                if isinstance(fitness_values_raw, np.ndarray):
                    fitness_values = fitness_values_raw.flatten()
                elif np.isscalar(fitness_values_raw): # Handles single numeric value
                    fitness_values = np.array([fitness_values_raw])
                else:
                    logging.warning(f"Unexpected type for fitness_values_raw in {file_name}, repeat {j+1}: {type(fitness_values_raw)}. Assigning None.")
                    best_fitness_per_repeat.append(None)
                    continue

                if fitness_values.size > 0:
                    try:
                        best_fitness_per_repeat.append(min(fitness_values))
                        logging.debug(f"Repeat {j+1}: Best fitness {min(fitness_values)}")
                    except ValueError: # Handles min() on empty sequence if somehow still occurs
                        logging.warning(f"ValueError: No valid fitness values to find minimum in {file_name}, repeat {j+1}. Assigning None.")
                        best_fitness_per_repeat.append(None)
                else:
                    logging.info(f"No fitness values found for {file_name}, repeat {j+1}. Assigning None.")
                    best_fitness_per_repeat.append(None)
            else:
                logging.warning(f"Problematic final_fitness structure or missing data for {file_name}, repeat {j+1}. Assigning None.")
                best_fitness_per_repeat.append(None)
        except Exception as e:
            logging.error(f"Unexpected error processing fitness for {file_name}, repeat {j+1}: {e}. Assigning None.")
            best_fitness_per_repeat.append(None)


    key = (year, func)
    grouped_data.setdefault(key, []).append({
        'selection_method': selection_methods,
        'best_fitnesses': best_fitness_per_repeat
    })

logging.info("Data processing loop finished.")

# Create Excel Workbook
wb = Workbook()
ws = wb.active
ws.title = "Experiment Results"

row_pointer = 1
logging.info("Starting Excel report generation...")

for (year, func), experiments in grouped_data.items():
    title = f"CEC{year}_F{func}"
    logging.info(f"Processing data for {title}")

    # Handle duplicate selection method names
    original_selection_methods = [m['selection_method'] for m in experiments]
    counts = {}
    unique_selection_methods = []
    for k_idx, method_name in enumerate(original_selection_methods):
        if method_name in counts:
            counts[method_name] += 1
            new_name = f"{method_name}_{counts[method_name]}"
            logging.info(f"Duplicate selection method '{method_name}' found for {title}. Renaming to '{new_name}'.")
            unique_selection_methods.append(new_name)
            experiments[k_idx]['selection_method'] = new_name # Update in the experiments list
        else:
            counts[method_name] = 1
            unique_selection_methods.append(method_name)
            # No need to update experiments[k_idx]['selection_method'] here as it's already unique

    if not experiments: # Should not happen if grouped_data has entries
        logging.warning(f"No experiments found for {title} after processing. Skipping.")
        continue
        
    # Ensure there's at least one experiment and it has 'best_fitnesses'
    if not experiments[0].get('best_fitnesses'):
        logging.warning(f"No 'best_fitnesses' found for the first experiment in {title}. Skipping this group.")
        continue

    num_repeats = len(experiments[0]['best_fitnesses'])
    if num_repeats == 0:
        logging.warning(f"Zero repeats found for {title} based on first experiment. Skipping this group.")
        continue


    # Create DataFrame
    columns = [m['selection_method'] for m in experiments] # Use updated names
    df = pd.DataFrame(index=[f"Repeat {i+1}" for i in range(num_repeats)])



    # Create DataFrame
    columns = [f"{m['selection_method']}" for m in experiments]
    df = pd.DataFrame(index=[f"Repeat {i+1}" for i in range(num_repeats)])
    
    
    for e in experiments:
        df[e["selection_method"]] = e["best_fitnesses"]

    # Merge header and write title
    start_col = 1
    end_col = start_col + df.shape[1]
    ws.merge_cells(start_row=row_pointer, start_column=start_col, end_row=row_pointer, end_column=end_col)
    cell = ws.cell(row=row_pointer, column=start_col)
    cell.value = title
    cell.alignment = Alignment(horizontal='center')
    row_pointer += 1

    # Write DataFrame
    for r in dataframe_to_rows(df.reset_index(), index=False, header=True):
        for c_idx, value in enumerate(r, start=1):
            ws.cell(row=row_pointer, column=c_idx, value=value)
        row_pointer += 1

    row_pointer += 2  # spacing between tables

# Save workbook
try:
    wb.save(output_excel)
    logging.info(f"Excel report successfully saved to: {output_excel}")
except Exception as e:
    logging.error(f"Failed to save Excel report to {output_excel}: {e}")

logging.info("Script finished.")
